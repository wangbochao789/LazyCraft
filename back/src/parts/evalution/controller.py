# Copyright (c) 2025 SenseTime. All Rights Reserved.
# Author: LazyLLM Team,  https://github.com/LazyAGI/LazyLLM
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

import gzip
import json
import logging
import os
import random
import shutil
import tarfile
import tempfile
import zipfile
from datetime import datetime
from threading import Thread

import pandas as pd
from flask import (copy_current_request_context, current_app, request,
                   send_from_directory)
from flask_login import current_user
from flask_restful import reqparse
from openpyxl import Workbook
from openpyxl.styles import Alignment

from core.restful import Resource
from libs.feature_gate import require_internet_feature
from libs.filetools import FileTools
from libs.helper import build_response
from libs.login import login_required
from parts.logs import Action, LogService, Module
from parts.urls import api
from utils.util_database import db

from .service import Service


class UploadDataset(Resource):
    """数据集上传API，用于上传和解析评估数据集。

    Args:
        file: 上传的压缩文件。

    Returns:
        dict: 上传结果信息。

    Raises:
        ValueError: 当文件类型不支持或文件内容不符合要求时抛出异常。
    """

    def unzip_file(self, file_path):
        """解压文件。

        Args:
            file_path (str): 要解压的文件路径。

        Returns:
            None: 无返回值。

        Raises:
            Exception: 当解压失败时抛出异常。
        """
        if file_path.endswith(".zip"):
            with zipfile.ZipFile(file_path, "r") as zip_ref:
                zip_ref.extractall(os.path.dirname(file_path))
        elif file_path.endswith(".tar.gz"):
            # 先解压缩gzip部分
            with gzip.open(file_path, "rb") as gz_ref:
                tar_file_path = os.path.splitext(file_path)[0]
                with open(tar_file_path, "wb") as tar_file:
                    tar_file.write(gz_ref.read())
            # 再解压缩tar部分
            with tarfile.open(tar_file_path, "r") as tar_ref:
                tar_ref.extractall(os.path.dirname(tar_file_path))
                # 删除临时解压的tar文件
                os.remove(tar_file_path)
        # elif file_path.endswith('.tar'):
        #     with tarfile.open(file_path, 'r') as tar_ref:
        #         tar_ref.extractall(os.path.dirname(file_path))

    def unzip_and_move_files(self, file_path):
        """解压文件并将解压后的所有文件移动到与压缩包同级的目录，处理重名文件，并删除压缩包。

        Args:
            file_path (str): 要解压的文件路径。

        Returns:
            None: 无返回值。

        Raises:
            Exception: 当解压或移动文件失败时抛出异常。
        """
        # 获取压缩包所在目录（即与压缩包同级的目录）
        try:
            parent_dir = os.path.dirname(file_path)
            tmp_storage_dir = FileTools.create_temp_storage(
                current_user.id, "evaluation" + str(random.randint(1, 1000))
            )

            if file_path.endswith(".zip"):
                with zipfile.ZipFile(file_path, "r") as zip_ref:
                    zip_ref.extractall(tmp_storage_dir)
            elif file_path.endswith(".tar"):
                with tarfile.open(file_path, "r") as tar_ref:
                    tar_ref.extractall(tmp_storage_dir)
            elif file_path.endswith(".tar.gz"):
                # 先解压缩gzip部分
                with gzip.open(file_path, "rb") as gz_ref:
                    tar_file_path = os.path.splitext(file_path)[0]
                    with open(tar_file_path, "wb") as tar_file:
                        tar_file.write(gz_ref.read())
                # 再解压缩tar部分
                with tarfile.open(tar_file_path, "r") as tar_ref:
                    tar_ref.extractall(tmp_storage_dir)
                    # 删除临时解压的tar文件
                    os.remove(tar_file_path)

            # 获取解压后的文件夹路径（假设解压后只有一个文件夹，可根据实际情况调整）
            # unzipped_folder_path = os.path.join(os.path.dirname(file_path), os.path.basename(file_path).split('.')[0])

            for root, dirs, files in os.walk(tmp_storage_dir):
                for file in files:
                    source_file_path = os.path.join(root, file)
                    destination_file_path = os.path.join(parent_dir, file)

                    # 如果有重名文件，修改名称后再移动
                    if os.path.exists(destination_file_path):
                        base_name, ext = os.path.splitext(file)
                        counter = 1
                        new_name = f"{base_name}_{counter}{ext}"
                        while os.path.exists(os.path.join(parent_dir, new_name)):
                            counter += 1
                            new_name = f"{base_name}_{counter}{ext}"

                        destination_file_path = os.path.join(parent_dir, new_name)

                    shutil.move(source_file_path, destination_file_path)

            # 删除解压后的文件夹
            shutil.rmtree(tmp_storage_dir)

            # 删除压缩包
            os.remove(file_path)
        except Exception as e:
            shutil.rmtree(tmp_storage_dir)
            logging.error(f"解压文件时发生错误: {e}")
            raise

    def check_extensions(self, folder_path):
        """检查文件夹中的文件扩展名是否合法。

        Args:
            folder_path (str): 要检查的文件夹路径。

        Returns:
            None: 无返回值。

        Raises:
            ValueError: 当发现不合法的文件扩展名时抛出异常。
        """
        allowed_extensions = ["json", "csv", "xlsx"]
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                if not any(file.endswith(ext) for ext in allowed_extensions):
                    raise ValueError(
                        f"压缩文件存在不合法文件: {os.path.join(root, file)}"
                    )

    # 上传数据集文件
    @login_required
    def post(self):
        """上传评估数据集文件。

        Args:
            通过multipart/form-data传递参数：
                files: 要上传的文件列表。

        Returns:
            dict: 上传结果信息，包含数据集ID。

        Raises:
            Exception: 当文件上传、解压或处理失败时抛出异常。
        """
        try:
            # 检查文件类型是否符合要求
            allowed_extensions = ["json", "csv", "xlsx", "tar.gz", "zip"]
            # 检查是否有文件在请求中
            if "files" not in request.files:
                return build_response(message="没有选择文件", status=400)

            files = request.files.getlist("files")

            if not files:
                return build_response(message="没有选择文件(no files)", status=400)

            storage_dir = FileTools.create_temp_storage(
                current_user.id, "evaluation" + str(random.randint(1, 1000))
            )
            for file in files:
                if not any(file.filename.endswith(ext) for ext in allowed_extensions):
                    return build_response(message="存在不合法的文件类型", status=400)
                if file.filename.endswith(".zip") or file.filename.endswith(".tar.gz"):
                    self.check_compres_package(file, allowed_extensions)

                # 检查文件大小是否在1GB以内
                max_size = 1024 * 1024 * 1024
                size = FileTools.get_file_size(file)
                if size > max_size:
                    return build_response(message="文件超过大小限制(1GB)", status=400)
                # 保存文件
                file_path = os.path.join(storage_dir, file.filename)
                file.save(file_path)
                if file_path.endswith(".tar.gz") or file_path.endswith(".zip"):
                    self.unzip_and_move_files(file_path)
                    self.check_extensions(os.path.dirname(file_path))
            dataset_id = Service().upload_datasets(storage_dir)
            shutil.rmtree(storage_dir)  # 无论成功与否都执行删除该目录
            if dataset_id:
                ret = build_response(
                    message="上传文件成功",
                    result={"dataset_id": dataset_id},
                    status=0,
                    code=200,
                )
                return ret
            else:
                return build_response(message="没有文件成功上传", status=400)
        except Exception as e:
            logging.error(f"上传文件失败: {e}")
            db.session.rollback()
            if os.path.exists(storage_dir):
                shutil.rmtree(storage_dir)
            return build_response(
                message=f"上传文件 {file.filename} 失败：" + str(e), status=400
            )

    def check_compres_package(self, file, allowed_ext):
        """检查压缩包内的文件类型是否符合要求。

        Args:
            file: 压缩文件对象。
            allowed_ext (list): 允许的文件扩展名列表。

        Returns:
            None: 无返回值。

        Raises:
            ValueError: 当压缩包内包含不支持的文件类型时抛出异常。
        """
        # 将列表转换为元组，因为endswith方法只接受字符串或字符串元组
        allowed_ext_tuple = tuple(allowed_ext)
        if file.filename.endswith(".zip"):
            with zipfile.ZipFile(file) as zf:
                for name in zf.namelist():
                    if not name.endswith(allowed_ext_tuple) and not name.endswith("/"):
                        raise ValueError(f"压缩包内有不支持的文件: {name}")
        elif file.filename.endswith(".tar.gz"):
            file.seek(0)
            with tarfile.open(fileobj=file, mode="r:gz") as tf:
                for member in tf.getmembers():
                    if member.isfile():
                        if not member.name.endswith(allowed_ext_tuple):
                            raise ValueError(f"压缩包内有不支持的文件: {member.name}")


class TaskList(Resource):
    """评估任务列表API，用于获取评估任务的分页列表。

    Args:
        page (int): 页码，默认为1。
        page_size (int): 每页大小，默认为20。

    Returns:
        dict: 分页结果，包含任务列表和分页信息。
    """

    @login_required
    def get(self):
        """获取评估任务分页列表。

        Args:
            通过URL参数传递：
                page (int, optional): 页码，默认为1。
                page_size (int, optional): 每页大小，默认为20。

        Returns:
            dict: 分页结果，包含任务列表和分页信息。
        """
        try:
            # 获取分页参数
            page = request.args.get("page", 1, type=int)
            per_page = request.args.get("per_page", 10, type=int)
            keyword = request.args.get("keyword")
            qtype = request.args.get("qtype")

            # 查询所有任务，并按创建时间降序排列
            pagination = Service().query_task_list(page, per_page, keyword, qtype)
            tasks = pagination.items

            # 格式化任务数据
            tasks_data = []
            for task in tasks:
                tasks_data.append(
                    {
                        "id": task.id,
                        "name": task.name,
                        "model_name": task.model_name,
                        "evaluation_method": task.evaluation_method,
                        "process": task.process,
                        "status": task.status,
                        "status_zh": task.status_zh,
                        "created_time": task.created_time,
                        "creator": task.username,
                    }
                )

            # 返回分页数据
            return build_response(
                result={
                    "tasks": tasks_data,
                    "total": pagination.total,
                    "pages": pagination.pages,
                    "current_page": pagination.page,
                    "per_page": pagination.per_page,
                }
            )

        except Exception as e:
            return build_response(message="获取任务失败：" + str(e), status=500)


class DeleteTask(Resource):
    """删除评估任务API，用于删除指定的评估任务。

    Args:
        task_id (str): 评估任务ID。

    Returns:
        dict: 删除结果信息。

    Raises:
        ValueError: 当任务ID为空或任务不存在时抛出异常。
    """

    @login_required
    def post(self, task_id):
        """删除指定的评估任务。

        Args:
            task_id (str): 评估任务ID。

        Returns:
            dict: 删除结果信息。

        Raises:
            ValueError: 当任务ID为空或任务不存在时抛出异常。
        """
        task = Service().get_task_by_id(task_id)
        self.check_can_admin()
        if not task:
            return build_response(message="任务不存在", status=404)
        self.check_user_del_perm(task.user_id)
        try:
            result = Service().delete_task(task)
            if not result:
                return build_response(message="任务不存在或已删除", status=404)
            LogService().add(
                Module.MODEL_EVALUATE,
                Action.EVALUATE_TASK_DELETE,
                task_method=task.evaluation_method_name,
                task_name=task.name,
            )
            return build_response(message="任务删除成功")
        except Exception as e:
            db.session.rollback()
            return build_response(message="删除任务失败：" + str(e), status=500)


class CreateTask(Resource):
    """创建评估任务API，用于创建新的评估任务。

    Args:
        name (str): 任务名称。
        description (str, optional): 任务描述。
        dataset_ids (list): 数据集ID列表。
        model_ids (list): 模型ID列表。
        evaluation_dimensions (list): 评估维度列表。

    Returns:
        dict: 创建结果信息，包含任务ID。

    Raises:
        ValueError: 当必要参数缺失时抛出异常。
    """

    @login_required
    def post(self):
        """创建新的评估任务。

        Args:
            通过JSON请求体传递参数：
                name (str): 任务名称。
                description (str, optional): 任务描述。
                dataset_ids (list): 数据集ID列表。
                model_ids (list): 模型ID列表。
                evaluation_dimensions (list): 评估维度列表。

        Returns:
            dict: 创建结果信息，包含任务ID。

        Raises:
            ValueError: 当必要参数缺失时抛出异常。
        """
        self.check_can_write()
        try:
            # 获取请求数据
            data = request.get_json()
            if (
                data.get("dataset_id", []) == []
                or type(data["dataset_id"]) is not list
                or data["dataset_id"] == []
            ):
                return build_response(message="请选择数据集", status=400)
            if data.get("task_name", "") == "":
                return build_response(message="请输入任务名称", status=400)
            if data.get("evaluation_method", "") == "":
                return build_response(message="请选择测评类型", status=400)
            if not data.get("model_name", ""):
                return build_response(message="请选择测评模型", status=400)
            if data.get("evaluation_method") == "ai" and (
                not data.get("ai_evaluator_name") or not data.get("prompt")
            ):
                return build_response(
                    message="请填写ai评测参数(测评器或prompt)", status=400
                )
            if data.get("prompt", ""):
                required_substrings = [
                    "{scene}",
                    "{scene_descrp}",
                    "{standard}",
                    "{instruction}",
                    "{output}",
                    "{response}",
                ]
                for substring in required_substrings:
                    if substring not in data["prompt"]:
                        return build_response(
                            message=f"prompt中缺少 {substring}", status=400
                        )
            if (
                data.get("ai_evaluator_name", "")
                and data["model_name"] == data["ai_evaluator_name"]
            ):
                return build_response(
                    message="ai测评器与被测模型不能为同一个", status=400
                )
            if data["evaluation_type"] == "offline":
                dataset_id = Service().set_offline_dataset_id(data["dataset_id"])
            else:
                # 在线数据集后续处理
                dataset_id = 0
            # 创建 Task 实例
            if Service().check_task_name_notunique(
                data["task_name"], current_user.current_tenant_id
            ):
                return build_response(message="任务名称已存在", status=400)
            task_id = Service().create_task(dataset_id, data)

            # 创建每个维度的实例
            Service().create_dimension(data["dimensions"], task_id)

            # 在线数据集要查找出平台的数据集文件处理以及ai测评
            @copy_current_request_context
            def process_evaluation(dataset_ids, task_id):
                logging.info(f"start process_evaluation: {dataset_id}, {task_id}")
                app = current_app._get_current_object()
                with app.app_context():
                    Service().process_online_dataset(dataset_ids, task_id)

            thread = Thread(
                target=process_evaluation, args=(data["dataset_id"], task_id)
            )
            thread.start()
            LogService().add(
                Module.MODEL_EVALUATE,
                Action.CREATE_EVALUATE_TASK,
                task_method=(
                    "AI测评" if data["evaluation_method"] == "ai" else "人工测评"
                ),
                task_name=data["task_name"],
            )
            return build_response(message="任务创建成功")
        except Exception as e:
            db.session.rollback()
            return build_response(message="任务创建失败：" + str(e), status=500)


class EvaluationDimension(Resource):
    """评估维度API，用于获取指定任务的评估维度信息。

    Args:
        task_id (str): 评估任务ID。

    Returns:
        dict: 评估维度信息。

    Raises:
        ValueError: 当任务不存在时抛出异常。
    """

    @login_required
    def get(self, task_id):
        """获取指定任务的评估维度信息。

        Args:
            task_id (str): 评估任务ID。

        Returns:
            dict: 评估维度信息。

        Raises:
            ValueError: 当任务不存在时抛出异常。
        """
        task = Service().get_task_by_id(task_id)
        self.check_can_read_object(task)
        dimensions = Service().get_evaluation_dimensions(task_id)
        dimensions_dict = [
            {
                "id": dimension.id,
                "dimension_name": dimension.dimension_name,
                "dimension_description": dimension.dimension_description,
                "ai_base_score": dimension.ai_base_score,
                "options": [
                    {
                        "id": option.id,
                        "option_description": option.option_description,
                        "value": option.value,
                    }
                    for option in dimension.options
                ],
            }
            for dimension in dimensions
        ]
        return build_response(result=dimensions_dict)


class Evaluate(Resource):
    """评估执行API，用于执行评估任务。

    Args:
        task_id (str): 评估任务ID。
        evaluation_data (list): 评估数据列表。

    Returns:
        dict: 评估结果信息。

    Raises:
        ValueError: 当必要参数缺失或任务不存在时抛出异常。
    """

    @login_required
    @require_internet_feature("模型评测")
    def post(self):
        """执行评估任务。

        Args:
            通过JSON请求体传递参数：
                task_id (str): 评估任务ID。
                evaluation_data (list): 评估数据列表。

        Returns:
            dict: 评估结果信息。

        Raises:
            ValueError: 当必要参数缺失或任务不存在时抛出异常。
        """
        try:
            data = request.json
            task_id = data.get("task_id")
            data_id = data.get("data_id")
            evaluations = data.get("evaluations")

            task = Service().get_task_by_id(task_id)
            self.check_can_admin_object(task)

            pre_completed = task.completed
            for eval_data in evaluations:
                Service().save_evaluation_scores(
                    task_id,
                    data_id,
                    eval_data["dimension_id"],
                    eval_data.get("option_select_id", 0),
                    eval_data.get("score", 0),
                )
            print(task.status)
            print(task.completed)
            print(task.total)
            if task.completed == 1 and task.evaluation_method == "manual":
                LogService().add(
                    Module.MODEL_EVALUATE,
                    Action.EVALUATE_TASK_START,
                    task_name=task.name,
                )

            if (task.total - pre_completed) == 1:
                LogService().add(
                    Module.MODEL_EVALUATE,
                    Action.EVALUATE_TASK_FINISH,
                    task_method=task.evaluation_method_name,
                    task_name=task.name,
                    completed=task.completed,
                    total=task.total,
                )

            return build_response(message="提交成功")
        except Exception as e:
            db.session.rollback()
            return build_response(message="提交失败：" + str(e), status=500)


class TaskInfo(Resource):
    """任务信息API，用于获取指定评估任务的详细信息。

    Args:
        task_id (str): 评估任务ID。

    Returns:
        dict: 任务详细信息。

    Raises:
        ValueError: 当任务不存在时抛出异常。
    """

    @login_required
    def get(self, task_id):
        """获取指定评估任务的详细信息。

        Args:
            task_id (str): 评估任务ID。

        Returns:
            dict: 任务详细信息。

        Raises:
            ValueError: 当任务不存在时抛出异常。
        """
        # 查询任务的基本信息
        task = Service().get_task_by_id(task_id)
        if not task:
            return build_response(message="任务不存在", status=404)
        self.check_can_read_object(task)

        # 任务的基本信息
        task_info = {
            "id": task.id,
            "name": task.name,
            "model_name": task.model_name,
            "evaluation_method": task.evaluation_method,
            "created_time": task.created_time,
            "username": task.username,
            "status": task.status,
            "status_zh": task.status_zh,
            "ai_eva_success": task.ai_eva_success,
            "ai_eva_fail": task.ai_eva_fail,
            "ai_evaluator_name": task.ai_evaluator_name,
            "process": task.process,
        }

        # 返回结果
        return build_response(result={"task_info": task_info})


class FirstUnScoreData(Resource):
    """首个未评分数据API，用于获取指定任务中第一个未评分的数据。

    Args:
        task_id (str): 评估任务ID。

    Returns:
        dict: 首个未评分的数据信息。

    Raises:
        ValueError: 当任务不存在时抛出异常。
    """

    @login_required
    def get(self, task_id):
        """获取指定任务中第一个未评分的数据。

        Args:
            task_id (str): 评估任务ID。

        Returns:
            dict: 首个未评分的数据信息。

        Raises:
            ValueError: 当任务不存在时抛出异常。
        """
        # 查询任务的基本信息
        task = Service().get_task_by_id(task_id)
        if not task:
            return build_response(message="任务不存在", status=404)
        first_unscored_data = Service().get_first_unscored_data(
            task_id=task_id, dataset_id=task.dataset_id
        )
        # 数据集信息
        if first_unscored_data:
            unscored_data_info = {
                "id": first_unscored_data.id,
                "input": first_unscored_data.input,
                "output": first_unscored_data.output,
                "response": first_unscored_data.response,
            }
        else:
            unscored_data_info = None
        return build_response(result={"unscored_data_info": unscored_data_info})


class EvaluationDataPaginator(Resource):
    """评估数据分页API，用于获取评估数据的分页列表。

    Args:
        task_id (str): 评估任务ID。
        page (int): 页码，默认为1。
        page_size (int): 每页大小，默认为20。

    Returns:
        dict: 分页结果，包含评估数据列表和分页信息。
    """

    def get(self, task_id):
        """获取评估数据分页列表。

        Args:
            task_id (str): 评估任务ID。
            通过URL参数传递：
                page (int, optional): 页码，默认为1。
                page_size (int, optional): 每页大小，默认为20。

        Returns:
            dict: 分页结果，包含评估数据列表和分页信息。
        """
        parser = reqparse.RequestParser()
        parser.add_argument(
            "page", type=int, required=False, location="args", help="Page number"
        )
        args = parser.parse_args()

        # 获取页码，默认为 None
        page = args.get("page", None)
        # 选项,报告查看
        option_select_id = args.get("option_select_id", None)
        task = Service().get_task_by_id(task_id)
        self.check_can_read_object(task)

        try:
            result = Service().get_evaluation_data(task, page, option_select_id)
            return build_response(result=result)
        except Exception as e:
            return build_response(message="获取数据失败：" + str(e), status=500)


# class EvluationDataset11(Resource):


class EvluationModel(Resource):
    """评估模型API，用于获取可用的评估模型列表。

    Returns:
        dict: 评估模型列表。
    """

    @login_required
    def get(self):
        """获取可用的评估模型列表。

        Returns:
            dict: 评估模型列表。
        """
        return build_response(result=Service().get_all_model())


class EvluationOnlineData(Resource):
    """评估在线数据API，用于获取在线评估数据。

    Returns:
        dict: 在线评估数据列表。
    """

    @login_required
    def get(self):
        """获取在线评估数据。

        Returns:
            dict: 在线评估数据列表。
        """
        return build_response(result=Service().get_all_dataset())


class EvaluationSummaryAPI(Resource):
    """评估总结API，用于获取指定任务的评估总结信息。

    Args:
        task_id (str): 评估任务ID。

    Returns:
        dict: 评估总结信息。

    Raises:
        ValueError: 当任务不存在时抛出异常。
    """

    @login_required
    def get(self, task_id):
        """获取指定任务的评估总结信息。

        Args:
            task_id (str): 评估任务ID。

        Returns:
            dict: 评估总结信息。

        Raises:
            ValueError: 当任务不存在时抛出异常。
        """
        try:
            # 获取任务信息
            task = Service().get_task_by_id(task_id)
            if not task:
                return build_response(message="任务不存在", status=404)

            self.check_can_read_object(task)
            return build_response(result=self.get_summary(task))
        except Exception as e:
            logging.error(f"EvaluationSummaryAPI.get error: {str(e)}")
            return build_response(message=f"获取评测总结失败: {str(e)}", status=500)

    def get_summary(self, task):
        """获取任务的评估总结。

        Args:
            task: 评估任务对象。

        Returns:
            dict: 评估总结信息。
        """
        summary_data = []
        logging.info(f"EvaluationSummaryAPI.get_summary: {task.id}")
        # 遍历每个维度
        dimensions = Service().get_evaluation_dimensions(task.id)
        dimension_scores = Service().dimension_score_for_task(task.id)
        for dimension in dimensions:
            dscore = dimension_scores.get(dimension.id, {"total": 0, "count": 0})
            dimension_summary = {
                "dimension_name": dimension.dimension_name,
                "average_score": 0,
                "std_dev": 0,
                "total_score": 0,
                "indicators": [],
            }

            total_score = 0
            score_list = []
            if task.evaluation_method == "ai":
                logging.info(
                    f"EvaluationSummaryAPI.get_summary 开始计算ai标准差: {dscore}"
                )
                # 获取每个维度的总分
                dimension_summary["total_score"] = dscore.get("total", 0) or 0
                count = dscore.get("count", 0) or 0
                dimension_summary["average_score"] = round(
                    0 if count == 0 else dimension_summary["total_score"] / count, 2
                )

                # 修复标准差计算 - 需要获取所有评分的详细信息来计算标准差
                import math

                try:
                    # 获取该维度的所有评分
                    from parts.evalution.model import EvaluationScore

                    scores = (
                        db.session.query(EvaluationScore.score)
                        .filter(
                            EvaluationScore.task_id == task.id,
                            EvaluationScore.dimension_id == dimension.id,
                        )
                        .all()
                    )

                    if scores and len(scores) > 1:
                        score_values = [int(s[0]) for s in scores if s[0] is not None]
                        if score_values:
                            mean = sum(score_values) / len(score_values)
                            variance = sum((x - mean) ** 2 for x in score_values) / (
                                len(score_values) - 1
                            )
                            dimension_summary["std_dev"] = round(math.sqrt(variance), 2)
                        else:
                            dimension_summary["std_dev"] = 0
                    else:
                        dimension_summary["std_dev"] = 0
                except Exception as e:
                    logging.error(
                        f"EvaluationSummaryAPI.get_summary 计算标准差失败: {str(e)}"
                    )
                    # 如果计算标准差失败，设置为0
                    dimension_summary["std_dev"] = 0
                logging.info(
                    f"EvaluationSummaryAPI.get_summary 计算标准差完成: {dimension_summary}"
                )

            else:
                # 遍历每个维度的指标
                dimension_options = Service().get_evaluation_dimensions_option(
                    dimension.id
                )
                logging.info(
                    f"EvaluationSummaryAPI.get_summary 人工测评维度选项: {len(dimension_options)}"
                )
                for option in dimension_options:
                    try:
                        option_count = Service().count_evaluation_option(
                            task.id, option.id
                        )
                        logging.info(
                            f"EvaluationSummaryAPI.get_summary 选项{option.id}计数: {option_count}"
                        )
                        # 计算该指标的总得分
                        option_total_score = option_count * option.value
                        total_score += option_total_score
                        score_list.append(option_total_score)
                        # 添加指标数据
                        dimension_summary["indicators"].append(
                            {
                                "option_id": option.id,
                                "name": option.option_description,
                                "score": option.value,
                                "total_score": option_total_score,
                                "percentage": 0,  # 稍后计算占比
                            }
                        )
                    except Exception as e:
                        logging.error(
                            f"EvaluationSummaryAPI.get_summary 处理选项{option.id}失败: {str(e)}"
                        )
                        # 如果处理某个选项失败，跳过该选项
                        continue

                # 计算平均分、标准差和各个指标占比
                if score_list:
                    dimension_summary["average_score"] = round(
                        sum(score_list) / len(score_list), 2
                    )
                    dimension_summary["std_dev"] = round(
                        (
                            sum(
                                (x - dimension_summary["average_score"]) ** 2
                                for x in score_list
                            )
                            / len(score_list)
                        )
                        ** 0.5,
                        2,
                    )
                dimension_summary["total_score"] = total_score

                for indicator in dimension_summary["indicators"]:
                    if total_score > 0:
                        indicator["percentage"] = (
                            f"{(indicator['total_score'] / total_score) * 100:.1f}%"
                        )

            summary_data.append(dimension_summary)
        # 返回结果
        return {
            "task_name": task.name,
            "created_by": task.username,
            "evaluation_method": task.evaluation_method,
            "progress": task.process,
            "dimensions": summary_data,
        }


class DownloadReportExcel(Resource):
    """下载评估报告Excel API，用于下载评估任务的Excel报告。

    Args:
        task_id (str): 评估任务ID。
        token (str, optional): 认证令牌。

    Returns:
        Response: Excel文件下载响应。

    Raises:
        ValueError: 当任务不存在或认证失败时抛出异常。
    """

    def check_login(self, token):
        """检查用户登录状态。

        Args:
            token (str): 认证令牌。

        Returns:
            bool: 登录是否有效。

        Raises:
            ValueError: 当认证失败时抛出异常。
        """
        import jwt

        if not token:
            return False
        try:
            payload = jwt.decode(token, current_app.secret_key, algorithms=["HS256"])
            user_id = payload.get("user_id")
            if user_id:
                return user_id
        except jwt.ExpiredSignatureError:
            return False
        except jwt.InvalidTokenError:
            return False
        return False

    def get(self, task_id):
        """下载评估任务的Excel报告。

        Args:
            task_id (str): 评估任务ID。
            通过URL参数传递：
                token (str, optional): 认证令牌。

        Returns:
            Response: Excel文件下载响应。

        Raises:
            ValueError: 当任务不存在或认证失败时抛出异常。
        """
        token = request.args.get("token")
        user_id = self.check_login(token)
        if not user_id:
            return current_app.login_manager.unauthorized()
        task = Service().get_task_by_id(task_id)
        summary = EvaluationSummaryAPI().get_summary(task)

        # 创建一个新的工作簿
        wb = Workbook()
        ws = wb.active
        # 设置表头数据
        data = [
            ["任务名：", summary["task_name"]],
            [
                "测评类型：",
                "人工测评" if summary["evaluation_method"] == "manual" else "AI评测",
            ],
            [],
            [],
        ]

        if summary["evaluation_method"] == "manual":
            evaluation_method_name = "人工测评"
            file_name_prefix = f'{summary["task_name"]}-人工评测'
            dimension_line = []
            option_head_line = []
            option_total_line = {}
            for dimension in summary["dimensions"]:
                dimension_line.append(dimension["dimension_name"])
                dimension_line.append(f'总分：{dimension["total_score"]}')
                dimension_line.append(f'平均分：{dimension["average_score"]}')
                dimension_line.append(f'标准差：{dimension["std_dev"]}')
                dimension_line.append("")
                dimension_line.append("")
                option_head_line.extend(["选项", "基础分", "得分", "占比", "", ""])
                # option_head_line.append(indicator.indicator_name)
                for index, indicator in enumerate(dimension["indicators"]):
                    option_line = option_total_line.get(index, [])
                    option_line.extend(
                        [
                            indicator["name"],
                            indicator["score"],
                            indicator["total_score"],
                            indicator["percentage"],
                            "",
                            "",
                        ]
                    )
                    option_total_line[index] = option_line
            data.append(dimension_line)
            data.append([])
            data.append(option_head_line)
            for option_line in option_total_line.values():
                data.append(option_line)
        else:
            evaluation_method_name = "AI评测"
            file_name_prefix = f'{summary["task_name"]}-AI评测'
            data.append(["维度", "对应分值", "平均分", "标准差"])
            for dimension in summary["dimensions"]:
                data.append(
                    [
                        dimension["dimension_name"],
                        dimension["total_score"],
                        dimension["average_score"],
                        dimension["std_dev"],
                    ]
                )

        # 将表头数据写入Excel
        for row_index, row_data in enumerate(data, start=1):
            for col_index, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=cell_value)
                cell.alignment = Alignment(horizontal="left", vertical="center")

        with tempfile.TemporaryDirectory() as tmpdirname:
            filename = "{}-{}.xlsx".format(
                file_name_prefix, datetime.now().strftime("%Y%m%d%H%M%S")
            )
            save_path = os.path.join(tmpdirname, filename)
            # 将Excel数据保存到文件
            wb.save(save_path)
            LogService().add(
                Module.MODEL_EVALUATE,
                Action.DOWNLOAD_EVALUATION_REPORT,
                user_id=user_id,
                task_method=evaluation_method_name,
                task_name=summary["task_name"],
            )

            return send_from_directory(tmpdirname, filename, as_attachment=True)


class DownloadDatasetTpl(Resource):
    """下载数据集模板API，用于下载评估数据集的模板文件。

    Args:
        template_type (str): 模板类型。

    Returns:
        Response: 模板文件下载响应。

    Raises:
        ValueError: 当模板类型不支持时抛出异常。
    """

    def get(self, template_type):
        """下载指定类型的评估数据集模板。

        Args:
            template_type (str): 模板类型。

        Returns:
            Response: 模板文件下载响应。

        Raises:
            ValueError: 当模板类型不支持时抛出异常。
        """
        data_tmplate_json = """[
                                {"instruction": "写一句双桶洗衣机的宣传语",
                                "output": "“双桶设计，洗衣更净一步，享受洁净生活新境界！”",
                                "response": "“分洗双桶，洁净加倍，让衣物享受专属SPA。”"},
                                {"instruction": "写一句AI Chatbot的宣传语",
                                "output": "智能对话，一触即达,让沟通无界限。",
                                "response": "你的智能对话伙伴，随时随地，懂你所需。"}
                            ]"""
        if template_type == "xlsx":
            # 定义数据
            data_tmplate = json.loads(data_tmplate_json)
            data = [
                [item["instruction"], item["output"], item["response"]]
                for item in data_tmplate
            ]

            # 创建工作簿和工作表
            wb = Workbook()
            ws = wb.active

            # 写入表头
            ws.append(["instruction", "output", "response"])

            # 写入数据
            for row in data:
                ws.append(row)
            with tempfile.TemporaryDirectory() as tmpdirname:
                filename = "dataset_template.xlsx"
                save_path = os.path.join(tmpdirname, filename)
                # 将Excel数据保存到文件
                wb.save(save_path)
                return send_from_directory(tmpdirname, filename, as_attachment=True)
        elif template_type == "csv":
            # 定义数据
            data_tmplate = json.loads(data_tmplate_json)
            data = {
                "instruction": [item["instruction"] for item in data_tmplate],
                "output": [item["output"] for item in data_tmplate],
                "response": [item["response"] for item in data_tmplate],
            }
            # 创建DataFrame
            df = pd.DataFrame(data)
            with tempfile.TemporaryDirectory() as tmpdirname:
                filename = "dataset_template.csv"
                save_path = os.path.join(tmpdirname, filename)
                df.to_csv(save_path, index=False)
                return send_from_directory(tmpdirname, filename, as_attachment=True)
        else:
            with tempfile.TemporaryDirectory() as tmpdirname:
                filename = "dataset_template.json"
                save_path = os.path.join(tmpdirname, filename)
                data = json.loads(data_tmplate_json)
                with open(save_path, "w") as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)
                return send_from_directory(tmpdirname, filename, as_attachment=True)


api.add_resource(UploadDataset, "/model_evalution/upload_dataset")
api.add_resource(CreateTask, "/model_evalution/create_task")
api.add_resource(
    EvaluationDimension, "/model_evalution/get_evaluation_dimensions/<int:task_id>"
)
# api.add_resource(FirstUnScoreData, '/model_evalution/first_unscored_data') #废弃
api.add_resource(Evaluate, "/model_evalution/evaluate_save")
api.add_resource(TaskInfo, "/model_evalution/task_info/<int:task_id>")
api.add_resource(TaskList, "/model_evalution/list")
api.add_resource(DeleteTask, "/model_evalution/delete_task/<int:task_id>")
api.add_resource(
    EvaluationDataPaginator, "/model_evalution/evaluation_data/<int:task_id>"
)
api.add_resource(EvluationModel, "/model_evalution/all_model")
api.add_resource(EvluationOnlineData, "/model_evalution/all_online_datasets")

api.add_resource(
    EvaluationSummaryAPI, "/model_evalution/evaluation_summary/<int:task_id>"
)

api.add_resource(
    DownloadReportExcel, "/model_evalution/evaluation_summary_download/<int:task_id>"
)
api.add_resource(
    DownloadDatasetTpl,
    "/model_evalution/evaluation_datasettpl_download/<string:template_type>",
)
